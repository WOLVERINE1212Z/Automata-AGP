"""
EXTRACTOR DE TABLAS - PLANOS TÉCNICOS AGP v2.1
===============================================
Lee archivos .dwg/.dxf de la ruta de red,
extrae todas las tablas/cajones y genera:
  1. JSON para la interfaz web
  2. Excel con todos los datos organizados

MEJORAS v2.1:
- Mejor logging y feedback en tiempo real
- Modo verbose para debugging
- Optimizado para muchos archivos
- Progress bar
aishhhh escribo muy aspero 
"""

import os
import sys
import json
import argparse
import io
import time
import threading
from pathlib import Path
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
from functools import partial

# Forzar UTF-8 en Windows
if sys.platform == 'win32':
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except:
        pass

# Intentar importar dependencias
print("[1/4] Verificando dependencias...:)", flush=True)

try:
    import ezdxf
except ImportError:
    print("   Instalando ezdxf...", flush=True)
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'ezdxf', '--quiet'])
    import ezdxf

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("   Instalando openpyxl...", flush=True)
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

print("   ✓ Dependencias listas", flush=True)


# ─────────────────────────────────────────
#  PROGRESS BAR
# ─────────────────────────────────────────

class ProgressBar:
    def __init__(self, total, prefix='Progress', length=40):
        self.total = total
        self.prefix = prefix
        self.length = length
        self.current = 0
        self.lock = threading.Lock()
        self.start_time = time.time()
    
    def update(self, n=1):
        with self.lock:
            self.current += n
            self._draw()
    
    def _draw(self):
        if not sys.stdout.isatty():
            # No drawing in non-interactive mode
            return
            
        percent = self.current / self.total if self.total > 0 else 0
        filled = int(self.length * percent)
        bar = '█' * filled + '░' * (self.length - filled)
        
        elapsed = time.time() - self.start_time
        rate = self.current / elapsed if elapsed > 0 else 0
        
        sys.stdout.write(f'\r{self.prefix}: |{bar}| {self.current}/{self.total} ({rate:.1f}/s)')
        sys.stdout.flush()
        
        if self.current >= self.total:
            sys.stdout.write('\n')


# ─────────────────────────────────────────
#  LOGGING
# ─────────────────────────────────────────

def log(msg, verbose_only=False, verbose=False):
    """Log con soporte para modo verbose."""
    if verbose_only and not verbose:
        return
    print(msg, flush=True)


def log_step(msg):
    """Log de paso importante."""
    print(f"\n{'='*50}", flush=True)
    print(f"  {msg}", flush=True)
    print('='*50, flush=True)


# ─────────────────────────────────────────
#  UTILIDADES
# ─────────────────────────────────────────

def get_file_hash(filepath):
    """Genera hash rápido del archivo para cache."""
    try:
        stat = os.stat(filepath)
        return f"{stat.st_size}_{stat.st_mtime}"
    except:
        return ""


def safe_filename(name):
    """Limpia nombre de archivo para Excel."""
    return name[:50].replace('/', '-').replace('\\', '-').replace(':', '').replace('*', '').replace('?', '')


# ─────────────────────────────────────────
#  EXTRACCIÓN DE TABLAS DESDE DXF/DWG
# ─────────────────────────────────────────

def extract_tables_from_file(filepath):
    """
    Extrae TODA la información posible de un archivo DWG/DXF.
    Busca: tablas, texto, bloques con atributos, etc.
    """
    tables_data = []

    try:
        # Intentar abrir el archivo
        try:
            doc = ezdxf.readfile(filepath)
        except Exception as e:
            # Si falla con DWG, intentar como DXF o reportar
            error_msg = str(e)
            if 'not a' in error_msg.lower() or 'dwg' in error_msg.lower():
                return [{'error': 'Archivo DWG no legible (formato antiguo)', 'file': filepath, 'tipo': 'dwg_legacy'}]
            raise
        
        msp = doc.modelspace()
        
        # 1. Buscar entidades TABLE (tablas formales de AutoCAD)
        for entity in msp:
            if entity.dxftype() in ('ACAD_TABLE', 'TABLE'):
                table_info = parse_table_entity(entity)
                if table_info:
                    table_info['source'] = 'Tabla'
                    tables_data.append(table_info)

        # 2. Buscar bloques con atributos en el modelspace
        blocks_in_doc = extract_blocks_from_modelspace(msp)
        tables_data.extend(blocks_in_doc)

        # 3. Buscar bloques con atributos en las definiciones
        blocks_data = extract_from_blocks(doc)
        tables_data.extend(blocks_data)

        # 4. Si no hay nada, buscar texto agrupado
        if not tables_data:
            tables_data = extract_from_text_blocks_fast(msp)

    except Exception as e:
        return [{'error': str(e), 'file': filepath}]

    return tables_data


def extract_blocks_from_modelspace(msp):
    """Extrae bloques con atributos del modelspace."""
    blocks_data = []
    
    for entity in msp:
        if entity.dxftype() == 'INSERT':
            try:
                attrs = list(entity.attribs)
                if attrs:
                    rows = []
                    for attr in attrs:
                        try:
                            tag = attr.dxf.get('tag', '')
                            text = attr.dxf.get('text', '')
                            if tag and text:
                                rows.append({'campo': tag.strip(), 'valor': text.strip()})
                        except:
                            continue
                    
                    if rows:
                        block_name = entity.dxf.get('name', 'Unknown')
                        fields = {}
                        for r in rows:
                            fields[r['campo']] = r['valor']
                        
                        blocks_data.append({
                            'source': f'Referencia: {block_name}',
                            'fields': fields,
                            'rows': rows,
                            'num_campos': len(rows)
                        })
            except:
                continue
    
    return blocks_data


def extract_from_blocks(doc):
    """Extrae información de bloques con atributos en las definiciones."""
    blocks_data = []
    
    for block in doc.blocks:
        # Ignorar bloques del sistema
        if block.name.startswith('*'):
            continue
        
        # Buscar insert entities con atributos
        for entity in block:
            if entity.dxftype() == 'INSERT':
                # Es una referencia de bloque, buscar atributos
                try:
                    attrs = list(entity.attribs)
                    if attrs:
                        rows = []
                        for attr in attrs:
                            try:
                                tag = attr.dxf.get('tag', '')
                                text = attr.dxf.get('text', '')
                                if tag and text:
                                    rows.append({'campo': tag.strip(), 'valor': text.strip()})
                            except:
                                continue
                        
                        if rows:
                            fields = {}
                            for r in rows:
                                fields[r['campo']] = r['valor']
                            
                            blocks_data.append({
                                'source': f'Bloque: {block.name}',
                                'fields': fields,
                                'rows': rows,
                                'num_campos': len(rows)
                            })
                except:
                    continue
    
    return blocks_data


def parse_table_entity(entity):
    """Parsea una entidad TABLE de AutoCAD."""
    try:
        rows = []
        num_rows = entity.dxf.get('num_rows', 0)
        num_cols = entity.dxf.get('num_cols', 0)

        if num_rows == 0 or num_cols == 0:
            return None

        cells = []
        try:
            cells = list(entity.cells)
        except:
            pass

        if cells:
            row_dict = {}
            for i, cell in enumerate(cells):
                r = i // max(num_cols, 1)
                if r not in row_dict:
                    row_dict[r] = []
                try:
                    text = cell.dxf.get('text', '') or ''
                    row_dict[r].append(str(text).strip())
                except:
                    row_dict[r].append('')

            for r in sorted(row_dict.keys()):
                row = row_dict[r]
                if any(c.strip() for c in row):
                    rows.append(row)

        if rows:
            return build_table_from_rows(rows)

    except:
        pass
    return None


def extract_from_text_blocks_fast(msp):
    """Alternativa: agrupa entidades TEXT/MTEXT cercanas."""
    texts = []

    for entity in msp:
        if entity.dxftype() not in ('TEXT', 'MTEXT'):
            continue
            
        try:
            if entity.dxftype() == 'TEXT':
                content = entity.dxf.get('text', '')
                try:
                    x = entity.dxf.insert.x
                    y = entity.dxf.insert.y
                except:
                    continue
            else:
                try:
                    content = entity.plain_mtext()
                    x = entity.dxf.insert.x
                    y = entity.dxf.insert.y
                except:
                    continue

            content = str(content).strip()
            if content and len(content) > 0:
                texts.append({'text': content, 'x': x, 'y': y})
        except:
            continue

    if not texts:
        return []

    texts.sort(key=lambda t: (-t['y'], t['x']))
    return group_texts_into_tables(texts)


def group_texts_into_tables(texts):
    """Agrupa textos cercanos en tablas."""
    if not texts:
        return []

    tables = []
    Y_THRESHOLD = 20
    X_GAP_THRESHOLD = 400

    # Agrupar por filas
    rows = []
    current_row = [texts[0]]
    current_y = texts[0]['y']

    for t in texts[1:]:
        if abs(t['y'] - current_y) <= Y_THRESHOLD:
            current_row.append(t)
        else:
            rows.append(sorted(current_row, key=lambda x: x['x']))
            current_row = [t]
            current_y = t['y']
    rows.append(sorted(current_row, key=lambda x: x['x']))

    # Separar tablas por gaps grandes
    table_rows = []
    last_y = rows[0][0]['y'] if rows else 0

    for row in rows:
        row_y = row[0]['y']
        if abs(row_y - last_y) > Y_THRESHOLD * 4 and table_rows:
            tbl = build_table_from_rows([[c['text'] for c in r] for r in table_rows])
            if tbl:
                tables.append(tbl)
            table_rows = []
        table_rows.append(row)
        last_y = row_y

    if table_rows:
        tbl = build_table_from_rows([[c['text'] for c in r] for r in table_rows])
        if tbl:
            tables.append(tbl)

    return tables


def build_table_from_rows(rows):
    """Construye un dict estructurado desde filas de texto."""
    if not rows:
        return None

    fields = {}
    raw_rows = []

    for row in rows:
        row = [str(c).strip() for c in row if str(c).strip()]
        
        if len(row) >= 2:
            key = row[0]
            val = ' '.join(row[1:])
            if key:
                fields[key] = val
                raw_rows.append({'campo': key, 'valor': val})
        elif len(row) == 1:
            raw_rows.append({'campo': row[0], 'valor': ''})

    if not raw_rows:
        return None

    return {
        'fields': fields,
        'rows': raw_rows,
        'num_campos': len(raw_rows)
    }


# ─────────────────────────────────────────
#  PROCESAMIENTO DE ARCHIVOS
# ─────────────────────────────────────────

def process_single_file(args):
    """Procesa un solo archivo. Designed para multiprocessing."""
    filepath, vehicle_path, verbose = args
    
    filename = os.path.basename(filepath)
    rel_path = os.path.relpath(filepath, vehicle_path)
    subfolder = str(Path(rel_path).parent)
    if subfolder == '.':
        subfolder = 'Raíz'
    
    # Extraer tablas
    tables = extract_tables_from_file(filepath)
    
    return {
        'archivo': filename,
        'ruta_relativa': rel_path,
        'subcarpeta': subfolder,
        'tablas': tables,
        'num_tablas': len([t for t in tables if isinstance(t, dict) and 'error' not in t]),
        'procesado': datetime.now().isoformat()
    }


# ─────────────────────────────────────────
#  SCAN DE CARPETAS
# ─────────────────────────────────────────

def scan_vehicle_folder(vehicle_path, use_cache=True, num_workers=4, verbose=False):
    """
    Escanea una carpeta de vehículo y extrae tablas.
    OPTIMIZADO con multiprocesamiento.
    """
    vehicle_name = Path(vehicle_path).name
    files_data = []
    
    # Cargar cache si existe
    cache_file = Path(vehicle_path) / '.planos_cache.json'
    cache = set()
    if use_cache and cache_file.exists():
        try:
            cache = set(json.loads(cache_file.read_text(encoding='utf-8')))
            log(f"  ✓ Cache cargado: {len(cache)} archivos", verbose=verbose, verbose_only=True)
        except:
            cache = set()

    # Recolectar todos los archivos
    log("  📂 Escaneando carpetas...", verbose=verbose)
    all_files = []
    dirs_scanned = 0
    
    for root, dirs, files in os.walk(vehicle_path):
        dirs[:] = [d for d in dirs if not d.startswith('.') and not d.startswith('$')]
        dirs_scanned += len(dirs)
        
        for filename in files:
            if filename.lower().endswith(('.dwg', '.dxf')):
                filepath = os.path.join(root, filename)
                all_files.append(filepath)

    if not all_files:
        log(f"  ⚠️  No se encontraron archivos DWG/DXF en {vehicle_name}")
        return {
            'vehiculo': vehicle_name,
            'ruta': str(vehicle_path),
            'total_archivos': 0,
            'archivos': [],
            'fecha_extraccion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        }

    log(f"  📄 {len(all_files)} archivo(s) DWG/DXF encontrado(s) en {dirs_scanned} carpetas", verbose=verbose)
    
    # Filtrar archivos no cacheados
    files_to_process = []
    cached_count = 0
    
    for f in all_files:
        rel_path = os.path.relpath(f, vehicle_path)
        file_hash = get_file_hash(f)
        cache_key = f"{rel_path}_{file_hash}"
        
        if cache_key in cache:
            cached_count += 1
        else:
            files_to_process.append(f)
    
    if cached_count > 0:
        log(f"  💾 {cached_count} archivo(s) en cache (ya procesados)", verbose=verbose)

    # Procesar archivos nuevos
    processed = 0
    errors = 0
    
    if files_to_process:
        log(f"  ⚙️  Procesando {len(files_to_process)} archivo(s) con {num_workers} workers...", verbose=verbose)
        
        args_list = [(f, vehicle_path, verbose) for f in files_to_process]
        
        try:
            with ProcessPoolExecutor(max_workers=num_workers) as executor:
                futures = {executor.submit(process_single_file, arg): arg for arg in args_list}
                
                for future in as_completed(futures):
                    try:
                        result = future.result(timeout=90)  # 90 seg timeout
                        if result:
                            files_data.append(result)
                            processed += 1
                    except Exception as e:
                        errors += 1
                        arg = futures[future]
                        log(f"    ⚠ Error: {os.path.basename(arg[0])}", verbose=verbose)
        except Exception as e:
            log(f"  ⚠ Error en procesamiento paralelo: {e}", verbose=verbose)
            # Fallback secuencial
            for f in files_to_process:
                try:
                    result = process_single_file((f, vehicle_path, verbose))
                    if result:
                        files_data.append(result)
                        processed += 1
                except:
                    errors += 1
        
        # Guardar cache
        new_cache_entries = set()
        for f in files_data:
            try:
                full_path = os.path.join(vehicle_path, f['ruta_relativa'])
                file_hash = get_file_hash(full_path)
                new_cache_entries.add(f"{f['ruta_relativa']}_{file_hash}")
            except:
                pass
        
        all_cache = cache | new_cache_entries
        try:
            cache_file.write_text(json.dumps(list(all_cache), ensure_ascii=False), encoding='utf-8')
        except:
            pass

    # Archivos en cache - agregar con datos mínimos
    files_from_cache = len(all_files) - len(files_to_process)
    
    return {
        'vehiculo': vehicle_name,
        'ruta': str(vehicle_path),
        'total_archivos': len(all_files),
        'archivos_procesados': processed,
        'archivos_cache': files_from_cache,
        'archivos_error': errors,
        'carpetas': dirs_scanned,
        'archivos': files_data,
        'fecha_extraccion': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }


def scan_all_vehicles(base_path, use_cache=True, num_workers=4, verbose=False):
    """Escanea todas las carpetas de vehículos en la ruta base."""
    base = Path(base_path)
    vehicles = []

    if not base.exists():
        log(f"\n❌ ERROR: No se puede acceder a {base_path}")
        log("\n💡 Sugerencias:")
        log("   1. Verifica que la red esté conectada")
        log("   2. Mapea la unidad: net use Z: \\\\192.168.2.37\\ingenieria")
        log("   3. Usa: python extractor_planos.py --ruta \"Z:\\PRODUCCION\\AGP PLANOS TECNICOS\"")
        return []

    log(f"\n{'='*60}")
    log(f"  AGP - EXTRACTOR DE PLANOS TÉCNICOS v2.1")
    log(f"{'='*60}")
    log(f"📂 Ruta: {base_path}")
    log(f"⚙️  Workers: {num_workers}")
    log(f"💾 Cache: {'Sí' if use_cache else 'No'}")
    
    # Contar vehículos primero
    folders = sorted([f for f in base.iterdir() if f.is_dir() and not f.name.startswith('.')])
    log(f"📁 Vehículos encontrados: {len(folders)}")
    
    if not folders:
        log("\n⚠️  No se encontraron carpetas de vehículos")
        return []

    for folder in folders:
        log(f"\n{'─'*60}")
        log(f"📦 Procesando: {folder.name}")
        data = scan_vehicle_folder(str(folder), use_cache=use_cache, num_workers=num_workers, verbose=verbose)
        vehicles.append(data)
        
        archivos_ok = len([f for f in data['archivos'] if f['num_tablas'] > 0])
        log(f"  ✓ {data.get('archivos_procesados', 0)} nuevo(s), {data.get('archivos_cache', 0)} en cache")
        log(f"  📊 {archivos_ok} archivo(s) con tablas")

    return vehicles


# ─────────────────────────────────────────
#  EXPORTAR A EXCEL
# ─────────────────────────────────────────

def export_to_excel(vehicles_data, output_path):
    """Exporta todos los datos a un Excel bien formateado."""
    wb = openpyxl.Workbook()

    COLOR_HEADER = "1B2A4A"
    COLOR_VEHICLE = "2E5C9E"
    COLOR_FILE = "4A90D9"
    COLOR_TABLE_HEADER = "E8F0FB"
    COLOR_ALT = "F5F8FF"

    def header_style(cell, bg=COLOR_HEADER, size=11, bold=True):
        cell.font = Font(bold=bold, color="FFFFFF", size=size, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def border_all(cell):
        thin = Side(style='thin', color="CCCCCC")
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Hoja RESUMEN
    ws = wb.active
    ws.title = "RESUMEN"

    ws.merge_cells('A1:F1')
    ws['A1'].value = "AGP PLANOS TÉCNICOS — RESUMEN"
    header_style(ws['A1'], size=14)
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:F2')
    ws['A2'].value = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(italic=True, color="666666")
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ["Vehículo", "Subcarpeta", "Archivo", "N° Tablas", "Campos", "Estado"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        header_style(c, bg=COLOR_VEHICLE)
        border_all(c)

    row = 5
    for v in vehicles_data:
        for f in v['archivos']:
            total_campos = sum(t.get('num_campos', 0) for t in f['tablas'] if isinstance(t, dict) and 'error' not in t)
            has_error = any('error' in t for t in f['tablas'] if isinstance(t, dict))
            estado = "⚠ Error" if has_error else ("✓ OK" if f['num_tablas'] > 0 else "— Sin tablas")

            data_row = [v['vehiculo'], f['subcarpeta'], f['archivo'], f['num_tablas'], total_campos, estado]
            for col, val in enumerate(data_row, 1):
                c = ws.cell(row=row, column=col, value=val)
                c.fill = PatternFill("solid", fgColor=COLOR_ALT if row % 2 == 0 else "FFFFFF")
                c.alignment = Alignment(vertical="center")
                border_all(c)
            row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15

    # Hojas por vehículo
    for v in vehicles_data:
        safe_name = safe_filename(v['vehiculo'])
        ws = wb.create_sheet(title=safe_name)

        ws.merge_cells('A1:D1')
        ws['A1'].value = f"🚗 {v['vehiculo']}"
        header_style(ws['A1'], bg=COLOR_VEHICLE, size=13)
        
        ws.merge_cells('A2:D2')
        ws['A2'].value = f"Archivos: {v['total_archivos']} | Extraído: {v['fecha_extraccion']}"
        ws['A2'].font = Font(italic=True, size=9)

        cur_row = 4

        for f in v['archivos']:
            ws.merge_cells(f'A{cur_row}:D{cur_row}')
            ws[f'A{cur_row}'].value = f"📄 {f['archivo']} ({f['subcarpeta']})"
            header_style(ws[f'A{cur_row}'], bg=COLOR_FILE)
            cur_row += 1

            if not f['tablas']:
                ws[f'A{cur_row}'] = "Sin tablas"
                ws[f'A{cur_row}'].font = Font(italic=True, color="999999")
                cur_row += 2
                continue

            for t_idx, table in enumerate(f['tablas'], 1):
                if isinstance(table, dict) and 'error' in table:
                    ws[f'A{cur_row}'] = f"Error: {table['error']}"
                    ws[f'A{cur_row}'].font = Font(color="CC0000")
                    cur_row += 1
                    continue

                if not isinstance(table, dict) or 'rows' not in table:
                    continue

                ws[f'A{cur_row}'] = f"  Tabla #{t_idx}"
                ws[f'A{cur_row}'].fill = PatternFill("solid", fgColor=COLOR_TABLE_HEADER)
                ws[f'A{cur_row}'].font = Font(bold=True)
                cur_row += 1

                for col, h in enumerate(["Campo", "Valor"], 1):
                    c = ws.cell(row=cur_row, column=col, value=h)
                    c.font = Font(bold=True, color="FFFFFF")
                    c.fill = PatternFill("solid", fgColor="3A6BB5")
                    border_all(c)
                cur_row += 1

                for i, field_row in enumerate(table.get('rows', [])):
                    bg = COLOR_ALT if i % 2 == 0 else "FFFFFF"
                    c1 = ws.cell(row=cur_row, column=1, value=field_row.get('campo', ''))
                    c2 = ws.cell(row=cur_row, column=2, value=field_row.get('valor', ''))
                    c1.fill = PatternFill("solid", fgColor=bg)
                    c2.fill = PatternFill("solid", fgColor=bg)
                    border_all(c1)
                    border_all(c2)
                    cur_row += 1

                cur_row += 1

        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 35

    wb.save(output_path)
    print(f"\n✅ Excel guardado: {output_path}")


# ─────────────────────────────────────────
#  EXPORTAR JSON
# ─────────────────────────────────────────

def export_to_json(vehicles_data, output_path):
    """Exporta los datos a JSON."""
    output = {
        'generado': datetime.now().isoformat(),
        'total_vehiculos': len(vehicles_data),
        'vehiculos': vehicles_data,
        'version': '2.1'
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON guardado: {output_path}")


# ─────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────

def main():
    print("\n" + "="*60, flush=True)
    print("  AGP - INICIANDO EXTRACTOR...", flush=True)
    print("="*60 + "\n", flush=True)
    
    parser = argparse.ArgumentParser(
        description='Extractor de tablas de planos técnicos AGP v2.1',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ejemplos:
  python extractor_planos.py                                    # Todo
  python extractor_planos.py --vehiculo "FORD RANGER"           # Solo uno
  python extractor_planos.py --workers 8 --verbose              # Rápido con debug
  python extractor_planos.py --no-cache --limpiar-cache         # Reprocesar todo
        """
    )
    parser.add_argument('--ruta', default=r'\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS',
                        help='Ruta base de los planos')
    parser.add_argument('--salida', default='./output', help='Carpeta de salida')
    parser.add_argument('--vehiculo', default=None, help='Solo un vehículo')
    parser.add_argument('--workers', type=int, default=4, help='Procesos paralelos')
    parser.add_argument('--no-cache', action='store_true', help='Sin cache')
    parser.add_argument('--limpiar-cache', action='store_true', help='Limpiar cache')
    parser.add_argument('--verbose', '-v', action='store_true', help='Modo debug')

    args = parser.parse_args()

    num_workers = max(1, min(args.workers, 16))
    output_dir = Path(args.salida)
    output_dir.mkdir(parents=True, exist_ok=True)

    json_path = output_dir / 'datos_planos.json'
    excel_path = output_dir / 'planos_tecnicos.xlsx'

    base_path = args.ruta
    vehicles_to_process = []

    if args.vehiculo:
        vehicle_path = Path(base_path) / args.vehiculo
        if vehicle_path.exists():
            vehicles_to_process.append(vehicle_path)
        else:
            log(f"\n❌ No encontrado: {vehicle_path}")
            base = Path(base_path)
            if base.exists():
                log("\nVehículos disponibles:")
                for f in sorted(base.iterdir()):
                    if f.is_dir() and not f.name.startswith('.'):
                        log(f"  - {f.name}")
            sys.exit(1)
    else:
        base = Path(base_path)
        if not base.exists():
            log(f"\n❌ ERROR: La ruta no existe: {base_path}")
            log("\n💡 Solución: Mapea la unidad de red primero:")
            log("   net use Z: \\\\192.168.2.37\\ingenieria")
            sys.exit(1)
        vehicles_to_process = [f for f in base.iterdir() if f.is_dir() and not f.name.startswith('.')]

    if args.limpiar_cache:
        log("\n🧹 Limpiando cache...")
        for v in vehicles_to_process:
            cache_file = v / '.planos_cache.json'
            if cache_file.exists():
                cache_file.unlink()
        log("   ✓ Cache limpiado")

    use_cache = not args.no_cache
    all_vehicles_data = []
    start_time = time.time()

    for vehicle_folder in vehicles_to_process:
        data = scan_vehicle_folder(str(vehicle_folder), use_cache=use_cache, 
                                   num_workers=num_workers, verbose=args.verbose)
        all_vehicles_data.append(data)

    elapsed = time.time() - start_time

    # Exportar
    print(f"\n{'='*60}")
    print("  📤 EXPORTANDO DATOS...")
    print('='*60)
    
    export_to_json(all_vehicles_data, str(json_path))
    export_to_excel(all_vehicles_data, str(excel_path))

    # Resumen
    total_files = sum(v['total_archivos'] for v in all_vehicles_data)
    total_tables = sum(sum(f['num_tablas'] for f in v['archivos']) for v in all_vehicles_data)
    total_procesados = sum(v.get('archivos_procesados', 0) for v in all_vehicles_data)
    total_cache = sum(v.get('archivos_cache', 0) for v in all_vehicles_data)

    print(f"\n{'='*60}")
    print("  ✅ EXTRACCIÓN COMPLETA")
    print('='*60)
    print(f"⏱️  Tiempo: {elapsed:.1f} segundos")
    print(f"📦 Vehículos: {len(all_vehicles_data)}")
    print(f"📄 Archivos totales: {total_files}")
    print(f"   - Nuevos: {total_procesados}")
    print(f"   - Cache: {total_cache}")
    print(f"📋 Tablas extraídas: {total_tables}")
    print(f"💾 JSON: {json_path}")
    print(f"📊 Excel: {excel_path}")


if __name__ == '__main__':
    main()

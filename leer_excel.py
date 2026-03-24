import openpyxl

wb = openpyxl.load_workbook('test_tecnica.xlsx')
print('Hojas:', wb.sheetnames)

for hoja in wb.sheetnames:
    print(f"\n=== {hoja} ===")
    ws = wb[hoja]
    print(f"Filas: {ws.max_row}, Columnas: {ws.max_column}")
    
    for i, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
        if any(cell is not None for cell in row):
            print(row)

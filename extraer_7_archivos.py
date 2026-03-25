# EXTRAER LOS 7 ARCHIVOS 090 DE TOYOTA
import os
import json
import subprocess
import shutil
import tempfile
from datetime import datetime
from collections import defaultdict

SERVIDOR = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
TOYOTA = "TOYOTA"
ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"

CAMPOS = {"OFFSET": "OFFSET", "BN": "BN", "BN+D": "BN+D", "BN INT": "BN INT", 
          "BANDA NEGRA": "BN", "ACERO": "ACERO", "STEEL": "ACERO", "OFFSET PC": "OFFSET PC"}

import ezdxf

print("\n" + "="*60)
print("  EXTRAER 090 DE TOYOTA (7 archivos)")
print("="*60 + "\n")

# 1. Buscar
print("[1] Buscando archivos...")
ruta_toyota = os.path.join(SERVIDOR, TOYOTA)
archivos = []
for root, dirs, files in os.walk(ruta_toyota):
    if '_DXF' in root:
        continue
    for f in files:
        if f.lower().endswith('090.dwg'):
            archivos.append({'ruta': os.path.join(root, f), 'nombre': f, 'carpeta': os.path.basename(root)})

print(f"   Encontrados: {len(archivos)}\n")

# 2. Procesar cada archivo
temp_dir = tempfile.mkdtemp(prefix="dxf_")
resultados = []
exitosos = 0

for i, info in enumerate(archivos, 1):
    print(f"[{i}/7] {info['nombre']}")
    
    carpeta = os.path.dirname(info['ruta'])
    
    # ODA
    cmd = [ODA_EXE, carpeta, temp_dir, "ACAD2018", "DXF", "0", "0"]
    subprocess.run(cmd, capture_output=True, text=True)
    
    # Buscar DXF
    dxfs = [f for f in os.listdir(temp_dir) if f.endswith('.dxf')]
    
    if dxfs:
        dxf_path = os.path.join(temp_dir, dxfs[0])
        
        try:
            doc = ezdxf.readfile(dxf_path)
            msp = doc.modelspace()
            
            # Extraer
            attrs = defaultdict(list)
            for block in msp.query('INSERT'):
                if hasattr(block, 'attribs') and block.attribs:
                    try:
                        x = block.dxf.insert.x if hasattr(block.dxf, 'insert') else 0
                        y = block.dxf.insert.y if hasattr(block.dxf, 'insert') else 0
                    except:
                        x, y = 0, 0
                    
                    for attr in block.attribs:
                        tag = getattr(attr.dxf, 'tag', '').strip() if hasattr(attr.dxf, 'tag') else ''
                        txt = getattr(attr.dxf, 'text', '').strip() if hasattr(attr.dxf, 'text') else ''
                        
                        if tag and txt:
                            campo = CAMPOS.get(tag.upper(), tag)
                            attrs[(round(x/10)*10, round(y/10)*10)].append({'c': campo, 'v': txt})
            
            # Tablas
            tablas = []
            for pos in sorted(attrs.keys()):
                atrs = attrs[pos]
                if not atrs:
                    continue
                
                datos = {}
                for a in atrs:
                    if a['c'] in datos:
                        datos[a['c']] = datos[a['c']] + " | " + a['v']
                    else:
                        datos[a['c']] = a['v']
                
                if any(c in datos for c in CAMPOS.values()):
                    tablas.append({'pos': pos, 'datos': datos})
            
            if tablas:
                exitosos += 1
                print(f"   OK: {len(tablas)} tablas")
                for t in tablas[:1]:
                    for k, v in t['datos'].items():
                        print(f"      {k}: {v[:25]}")
            else:
                print(f"   Sin datos")
            
            resultados.append({'archivo': info['nombre'], 'carpeta': info['carpeta'], 'tablas': tablas,
                              'resumen': {k: v for t in tablas for k, v in t['datos'].items()}})
        except Exception as e:
            print(f"   Error: {e}")
            resultados.append({'archivo': info['nombre'], 'carpeta': info['carpeta'], 'error': str(e), 'tablas': []})
    else:
        print(f"   Error DXF")
        resultados.append({'archivo': info['nombre'], 'carpeta': info['carpeta'], 'error': 'No DXF', 'tablas': []})
    
    print()
    
    # Limpiar
    for f in os.listdir(temp_dir):
        if f.endswith('.dxf'):
            try:
                os.remove(os.path.join(temp_dir, f))
            except:
                pass

print("="*60)
print(f"RESUMEN: {len(resultados)} archivos, {exitosos} con datos\n")

# Guardar JSON
os.makedirs("output", exist_ok=True)
json.dump({
    'metadata': {'pieza': '090', 'carpeta': TOYOTA, 'total': len(resultados), 'exitosos': exitosos, 'fecha': datetime.now().isoformat()},
    'planos': resultados
}, open('output/090_toyota.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=2)

print("Guardado: output/090_toyota.json")

# Excel
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    wb = openpyxl.Workbook()
    por_carpeta = defaultdict(list)
    for p in resultados:
        por_carpeta[p['carpeta']].append(p)
    
    wb.remove(wb.active)
    
    for carpeta, planos in sorted(por_carpeta.items()):
        ws = wb.create_sheet(title=(carpeta[:31] if carpeta else "Raiz"))
        for col, h in enumerate(["Archivo", "Campo", "Valor"], 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="00A2D7", fill_type="solid")
        
        row = 2
        for p in planos:
            for t in p.get('tablas', []):
                for cmp, val in t.get('datos', {}).items():
                    ws.cell(row=row, column=1, value=p['archivo'])
                    ws.cell(row=row, column=2, value=cmp)
                    ws.cell(row=row, column=3, value=val)
                    row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
    
    wb.save('output/090_toyota.xlsx')
    print("Guardado: output/090_toyota.xlsx")
except Exception as e:
    print(f"Error Excel: {e}")

shutil.rmtree(temp_dir, ignore_errors=True)
print("\nFINALIZADO!")
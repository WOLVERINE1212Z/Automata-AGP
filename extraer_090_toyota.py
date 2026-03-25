#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
EXTRACTOR 090 TOYOTA - Extracción directa de campos técnicos
Busca campos técnicos específicos y extrae el valor a su derecha
"""
import os
import sys
import json
import subprocess
import shutil
import tempfile
import time
from datetime import datetime
from collections import defaultdict

# ============== CONFIGURACIÓN ==============
SERVIDOR = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
VEHICULO = "TOYOTA"
SUFIJO = "090.dwg"

ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"

# Ruta de imágenes (usa Z: si está mapeada, si no usa UNC)
RUTA_IMAGENES = r"Z:\PlanosSapProduccion"
if not os.path.exists(RUTA_IMAGENES):
    RUTA_IMAGENES = r"\\192.168.2.2\Sapfiles\PlanosSapProduccion"

import re

# Campos técnicos a buscar - SOLO palabras clave técnicas
CAMPOS_A_BUSCAR = [
    "OFFSET", "BN+D", "BN INT", "ACERO", "STEEL", "ESPESOR", 
    "LARGO", "ANCHO", "PESO", "MATERIAL", "BANDA", "TIPO", "MEDIDA"
]

import ezdxf


def buscar_imagenes_por_archivo(nombre_dwg):
    """Busca imágenes relacionadas con el archivo DWG"""
    imagenes = []
    try:
        nombre_base = os.path.splitext(nombre_dwg)[0].upper()
        
        # Extraer grupos de dígitos separados por no-dígitos
        # "1490 005 090" -> ["1490", "005", "090"]
        nums = re.split(r'[^0-9]+', nombre_base)
        nums = [n for n in nums if n]  # Filtrar vacíos
        
        if len(nums) >= 2:
            proyecto = nums[0]
            pieza = nums[-1]
            
            if os.path.exists(RUTA_IMAGENES):
                archivos_img = os.listdir(RUTA_IMAGENES)
                for img in archivos_img:
                    if not img.lower().endswith(('.jpg', '.jpeg', '.png')):
                        continue
                    
                    # Limpiar para comparación: quitar espacios y M inicial
                    img_clean = img.upper().replace(' ', '').replace('M', '')
                    
                    # Verificar coincidencia de proyecto + pieza
                    if proyecto in img_clean and pieza in img_clean:
                        img_path = os.path.join(RUTA_IMAGENES, img)
                        if img_path not in imagenes:
                            imagenes.append(img_path)
                            print(f"    -> IMG: {img}")
    except Exception as e:
        print(f"Error imágenes: {e}")
    return imagenes[:10]


def print_progress(msg, tipo="INFO"):
    hora = datetime.now().strftime("%H:%M:%S")
    print(f"[{hora}] [{tipo}] {msg}")
    sys.stdout.flush()


def buscar_archivos(servidor, vehiculo, sufijo):
    print_progress(f"Buscando archivos *{sufijo} en {vehiculo}...")
    ruta_vehiculo = os.path.join(servidor, vehiculo)
    archivos = []
    
    for root, dirs, files in os.walk(ruta_vehiculo):
        if '_DXF' in root:
            continue
        for f in files:
            if f.lower().endswith(sufijo.lower()):
                archivos.append({
                    'ruta': os.path.join(root, f),
                    'nombre': f,
                    'carpeta': os.path.basename(root)
                })
    
    print_progress(f"Encontrados {len(archivos)} archivos", "OK")
    return archivos


def convertir_dwg_a_dxf(dwg_path, temp_dir):
    carpeta = os.path.dirname(dwg_path)
    nombre = os.path.basename(dwg_path)
    
    cmd = [ODA_EXE, carpeta, temp_dir, "ACAD2018", "DXF", "0", "0"]
    subprocess.run(cmd, capture_output=True, text=True)
    
    nombre_dxf = os.path.splitext(nombre)[0] + ".dxf"
    dxf_path = os.path.join(temp_dir, nombre_dxf)
    
    if os.path.exists(dxf_path):
        return dxf_path
    
    dxfs = [f for f in os.listdir(temp_dir) if f.endswith('.dxf')]
    if dxfs:
        return os.path.join(temp_dir, dxfs[0])
    
    return None


def extraer_campos_tecnicos(dxf_path):
    """
    Busca campos técnicos específicos en el DXF y extrae el valor a la derecha
    """
    try:
        doc = ezdxf.readfile(dxf_path)
        msp = doc.modelspace()
        
        # Recolectar todos los textos con su posición
        textos = []
        
        # TEXT
        for text in msp.query('TEXT'):
            try:
                x = text.dxf.insert.x
                y = text.dxf.insert.y
                txt = text.dxf.text.strip()
                if txt:
                    textos.append({'x': x, 'y': y, 'texto': txt})
            except:
                pass
        
        # MTEXT
        for mtext in msp.query('MTEXT'):
            try:
                if hasattr(mtext.dxf, 'insert'):
                    x = mtext.dxf.insert.x
                    y = mtext.dxf.insert.y
                else:
                    x, y = 0, 0
                txt = mtext.text.strip() if hasattr(mtext, 'text') else ''
                if txt:
                    txt = txt.replace('\\P', ' ').replace('\\~', ' ').strip()
                    textos.append({'x': x, 'y': y, 'texto': txt})
            except:
                pass
        
        # ATTRIB
        for block_ref in msp.query('INSERT'):
            if hasattr(block_ref, 'attribs') and block_ref.attribs:
                try:
                    bx = block_ref.dxf.insert.x if hasattr(block_ref.dxf, 'insert') else 0
                    by = block_ref.dxf.insert.y if hasattr(block_ref.dxf, 'insert') else 0
                except:
                    bx, by = 0, 0
                
                for attr in block_ref.attribs:
                    try:
                        tag = getattr(attr.dxf, 'tag', '').strip() if hasattr(attr.dxf, 'tag') else ''
                        txt = getattr(attr.dxf, 'text', '').strip() if hasattr(attr.dxf, 'text') else ''
                        if tag and txt:
                            # El tag es el campo, el texto es el valor
                            # Buscar el tag como campo técnico
                            if any(tag.upper().startswith(c.upper()) for c in CAMPOS_A_BUSCAR):
                                textos.append({'x': bx, 'y': by, 'texto': tag})
                                textos.append({'x': bx + 10, 'y': by, 'texto': txt})
                    except:
                        pass
        
        if not textos:
            return []
        
        #BUSCAR PALABRAS CLAVE Y SUS VALORES (derecha Y abajo - formato tabla)
        datos_encontrados = {}
        
        for t in textos:
            texto_upper = t['texto'].upper().strip()
            
            # Verificar si es una palabra clave
            es_campo = False
            for kw in CAMPOS_A_BUSCAR:
                if kw.upper() in texto_upper or texto_upper.startswith(kw.upper()):
                    es_campo = True
                    break
            
            if es_campo:
                campo = t['texto'].strip()
                valores_encontrados = []
                
                # 1. Buscar valores a la DERECHA (misma fila)
                for otro in textos:
                    if abs(otro['y'] - t['y']) < 2 and otro['x'] > t['x']:
                        distancia = otro['x'] - t['x']
                        if distancia < 80:  # Mayor distancia para valores a la derecha
                            otro_upper = otro['texto'].upper()
                            es_otra_kw = any(kw.upper() in otro_upper or otro_upper.startswith(kw.upper()) for kw in CAMPOS_A_BUSCAR)
                            if not es_otra_kw and otro['texto'].strip():
                                valores_encontrados.append(otro['texto'].strip())
                
                # 2. Buscar valores ABAJO (misma columna, filas inferiores)
                # Ordenar textos por Y para encontrar los de abajo
                textos_debajo = [otro for otro in textos if otro['y'] < t['y'] - 2 and otro['y'] > t['y'] - 30]
                textos_debajo.sort(key=lambda x: x['y'], reverse=True)  # De más cerca a más lejos
                
                for otro in textos_debajo:
                    if abs(otro['x'] - t['x']) < 10:  # Misma columna (X aproximada)
                        otro_upper = otro['texto'].upper()
                        es_otra_kw = any(kw.upper() in otro_upper or otro_upper.startswith(kw.upper()) for kw in CAMPOS_A_BUSCAR)
                        if not es_otra_kw and otro['texto'].strip():
                            valores_encontrados.append(otro['texto'].strip())
                
                if valores_encontrados:
                    # Unir valores con |
                    valor_final = " | ".join(valores_encontrados)
                    if campo in datos_encontrados:
                        if valor_final not in datos_encontrados[campo]:
                            datos_encontrados[campo] = datos_encontrados[campo] + " | " + valor_final
                    else:
                        datos_encontrados[campo] = valor_final
        
        if datos_encontrados:
            return [{'atributos': datos_encontrados}]
        
        return []
        
    except Exception as e:
        print_progress(f"Error: {e}", "ERR")
        return []


def procesar_archivos(archivos, temp_dir):
    resultados = []
    exitosos = 0
    errores = 0
    
    total = len(archivos)
    print_progress(f"Procesando {total} archivos...")
    
    for i, info in enumerate(archivos, 1):
        print_progress(f"[{i}/{total}] {info['nombre']}", "PROC")
        
        dwg_path = info['ruta']
        dxf_path = convertir_dwg_a_dxf(dwg_path, temp_dir)
        
        # Buscar imágenes relacionadas
        imagenes = buscar_imagenes_por_archivo(info['nombre'])
        if imagenes:
            print_progress(f"  -> {len(imagenes)} imágenes encontradas", "IMG")
        
        if dxf_path and os.path.exists(dxf_path):
            tablas = extraer_campos_tecnicos(dxf_path)
            
            if tablas and tablas[0].get('atributos'):
                exitosos += 1
                resumen = tablas[0]['atributos']
                
                for campo, valor in list(resumen.items())[:3]:
                    print_progress(f"  -> {campo}: {valor}", "DAT")
                
                resultados.append({
                    'archivo': info['nombre'],
                    'carpeta': info['carpeta'],
                    'tablas': tablas,
                    'imagenes': imagenes,
                    'resumen': resumen
                })
            else:
                print_progress(f"  -> Sin datos técnicos", "WARN")
                errores += 1
                resultados.append({
                    'archivo': info['nombre'],
                    'carpeta': info['carpeta'],
                    'tablas': [],
                    'imagenes': imagenes,
                    'resumen': {}
                })
            
            try:
                os.remove(dxf_path)
            except:
                pass
        else:
            print_progress(f"  -> Error DXF", "ERR")
            errores += 1
            resultados.append({
                'archivo': info['nombre'],
                'carpeta': info['carpeta'],
                'error': 'DXF error',
                'tablas': [],
                'resumen': {}
            })
    
    return resultados, exitosos, errores


def guardar_json(resultados, exitosos, errores):
    os.makedirs("output", exist_ok=True)
    
    output_data = {
        'metadata': {
            'vehiculo': VEHICULO,
            'sufijo': SUFIJO,
            'total_archivos': len(resultados),
            'exitosos': exitosos,
            'errores': errores,
            'fecha_extraccion': datetime.now().isoformat()
        },
        'planos': resultados
    }
    
    json_path = f"output/{VEHICULO}_{SUFIJO.replace('.dwg','')}.json"
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    print_progress(f"JSON: {json_path}", "OK")
    return json_path


def guardar_excel(resultados):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Datos Extraídos"
        
        headers = ["Archivo", "Carpeta", "Campo", "Valor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="00A2D7", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        for p in resultados:
            archivo = p.get('archivo', '')
            carpeta = p.get('carpeta', '')
            tablas = p.get('tablas', [])
            
            for t in tablas:
                for campo, valor in t.get('atributos', {}).items():
                    ws.cell(row=row, column=1, value=archivo)
                    ws.cell(row=row, column=2, value=carpeta)
                    ws.cell(row=row, column=3, value=campo)
                    ws.cell(row=row, column=4, value=valor)
                    row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        
        excel_path = f"output/{VEHICULO}_{SUFIJO.replace('.dwg','')}.xlsx"
        wb.save(excel_path)
        
        print_progress(f"Excel: {excel_path}", "OK")
        return excel_path
        
    except Exception as e:
        print_progress(f"Error Excel: {e}", "ERR")
        return None


def main():
    print("\n" + "="*60)
    print("  EXTRACTOR 090 TOYOTA - Campos Técnicos")
    print("="*60 + "\n")
    
    inicio = time.time()
    temp_dir = tempfile.mkdtemp(prefix="dxf_")
    
    try:
        archivos = buscar_archivos(SERVIDOR, VEHICULO, SUFIJO)
        
        if not archivos:
            print_progress("No se encontraron archivos", "WARN")
            return
        
        resultados, exitosos, errores = procesar_archivos(archivos, temp_dir)
        json_path = guardar_json(resultados, exitosos, errores)
        excel_path = guardar_excel(resultados)
        
        print("\n" + "="*60)
        print("  RESUMEN")
        print("="*60)
        print(f"  Total:       {len(resultados)}")
        print(f"  Exitosos:   {exitosos}")
        print(f"  Errores:    {errores}")
        print(f"  JSON:       {json_path}")
        if excel_path:
            print(f"  Excel:      {excel_path}")
        
        duracion = time.time() - inicio
        print(f"\n  Tiempo:     {duracion:.1f}s")
        print("="*60 + "\n")
        
    finally:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except:
            pass


if __name__ == "__main__":
    main()
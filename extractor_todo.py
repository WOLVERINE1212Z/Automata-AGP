# -*- coding: utf-8 -*-
"""
EXTRACTOR COMPLETO - Procesa TODOS los archivos DWG del servidor
Mantiene estructura de carpetas exacta
"""

import os
import sys
import json
import subprocess
import shutil
import tempfile
from datetime import datetime

# Configurar UTF-8
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ================== CONFIGURACION ==================
SERVIDOR = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
SALIDA_JSON = r"output\datos_planos.json"
SALIDA_EXCEL = r"output\planos_tecnicos.xlsx"
CARPETA_DXF = r"\\192.168.2.37\ingenieria\PRODUCCION\_DXF_CONVERTIDOS"
ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"

# Importar ezdxf
try:
    import ezdxf
except ImportError:
    print("[ERROR] Instalar ezdxf: pip install ezdxf")
    sys.exit(1)

# ================== FUNCIONES ==================

def obtener_todos_dwg():
    """Busca TODOS los archivos DWG en el servidor"""
    print("[1] Buscando todos los archivos DWG...")
    archivos = []
    
    for root, dirs, files in os.walk(SERVIDOR):
        if '_DXF_CONVERTIDOS' in root:
            continue
            
        for file in files:
            if file.lower().endswith('.dwg'):
                ruta_completa = os.path.join(root, file)
                ruta_rel = ruta_completa.replace(SERVIDOR, '').strip('\\')
                archivos.append({
                    'ruta_completa': ruta_completa,
                    'ruta_relativa': ruta_rel,
                    'carpeta': os.path.dirname(ruta_rel),
                    'nombre': file
                })
    
    print(f"    Encontrados: {len(archivos)} archivos DWG")
    return archivos

def convertir_dwg(dwg_path, carpeta_salida):
    """Convierte un DWG a DXF usando ODA"""
    carpeta_entrada = os.path.dirname(dwg_path)
    nombre_dwg = os.path.basename(dwg_path)
    
    try:
        cmd = [
            ODA_EXE,
            carpeta_entrada,
            carpeta_salida,
            "ACAD2018",
            "DXF",
            "0",
            "0"
        ]
        
        resultado = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300
        )
        
        if resultado.returncode == 0:
            nombre_dxf = os.path.splitext(nombre_dwg)[0] + ".dxf"
            ruta_dxf = os.path.join(carpeta_salida, nombre_dxf)
            
            if os.path.exists(ruta_dxf):
                return ruta_dxf
            
            for f in os.listdir(carpeta_salida):
                if f.lower().endswith('.dxf'):
                    return os.path.join(carpeta_salida, f)
        
        return None
        
    except Exception as e:
        return None

def extraer_datos(ruta_dxf, info_archivo):
    """Extrae datos técnicos del DXF"""
    datos = {
        'archivo': info_archivo['nombre'],
        'ruta': info_archivo['ruta_relativa'],
        'carpeta': info_archivo['carpeta'],
        'tablas': [],
        'error': None
    }
    
    try:
        doc = ezdxf.readfile(ruta_dxf)
        msp = doc.modelspace()
        
        bloques = list(msp.query('INSERT'))
        
        for block_ref in bloques:
            if hasattr(block_ref, 'attribs') and block_ref.attribs:
                nombre_bloque = block_ref.dxf.name
                atributos = {}
                
                for attr in block_ref.attribs:
                    tag = getattr(attr.dxf, 'tag', '').strip() if hasattr(attr.dxf, 'tag') else ''
                    texto = getattr(attr.dxf, 'text', '').strip() if hasattr(attr.dxf, 'text') else ''
                    
                    if tag and texto and tag not in atributos:
                        atributos[tag] = texto
                
                if atributos:
                    existe = False
                    for t in datos['tablas']:
                        if t.get('atributos') == atributos:
                            existe = True
                            break
                    
                    if not existe:
                        datos['tablas'].append({
                            'bloque': nombre_bloque,
                            'atributos': atributos
                        })
        
    except Exception as e:
        datos['error'] = str(e)
    
    return datos

def main():
    print("="*60)
    print("EXTRACTOR COMPLETO - TODOS LOS ARCHIVOS")
    print("="*60)
    print(f"Servidor: {SERVIDOR}")
    print(f"Inicio: {datetime.now()}")
    print()
    print("IMPORTANTE: Los archivos DWG originales NO se tocan")
    print("Solo se convierten a DXF temporalmente")
    print()
    
    # Obtener todos los archivos
    archivos = obtener_todos_dwg()
    
    if not archivos:
        print("[ERROR] No hay archivos DWG")
        return
    
    # Carpeta temporal para DXF
    temp_dir = tempfile.mkdtemp(prefix="dxf_completo_")
    print(f"\n[2] Procesando {len(archivos)} archivos...")
    print(f"    Esto puede tardar varias horas...")
    print(f"    Carpeta temp: {temp_dir}")
    
    resultados = []
    exitosos = 0
    errores = 0
    
    for i, info in enumerate(archivos, 1):
        nombre = info['nombre']
        
        # Mostrar progreso cada 50 archivos
        if i % 50 == 0 or i == 1:
            print(f"  [{i}/{len(archivos)}] {info['carpeta'][:50]}... ", end="", flush=True)
        
        # Convertir
        dxf_path = convertir_dwg(info['ruta_completa'], temp_dir)
        
        if dxf_path and os.path.exists(dxf_path):
            datos = extraer_datos(dxf_path, info)
            
            if datos.get('tablas'):
                exitosos += 1
                if i % 50 == 0:
                    print(f"OK ({len(datos['tablas'])} tablas)")
            else:
                if i % 50 == 0:
                    print("Sin datos")
            
            resultados.append(datos)
        else:
            errores += 1
            if i % 50 == 0:
                print("Error")
            resultados.append({
                'archivo': nombre,
                'carpeta': info['carpeta'],
                'ruta': info['ruta_relativa'],
                'error': 'No se pudo convertir',
                'tablas': []
            })
        
        # Limpiar DXF de la carpeta temporal cada 100 archivos
        if i % 100 == 0:
            try:
                for f in os.listdir(temp_dir):
                    if f.endswith('.dxf'):
                        os.remove(os.path.join(temp_dir, f))
            except:
                pass
    
    # Resumen
    print(f"\n[3] Resumen:")
    print(f"    Procesados: {len(resultados)}")
    print(f"    Con tablas: {exitosos}")
    print(f"    Errores: {errores}")
    
    # Guardar JSON
    print(f"\n[4] Guardando JSON...")
    os.makedirs(os.path.dirname(SALIDA_JSON), exist_ok=True)
    
    datos_json = {
        'metadata': {
            'servidor': SERVIDOR,
            'total': len(resultados),
            'exitosos': exitosos,
            'errores': errores,
            'fecha': datetime.now().isoformat()
        },
        'planos': resultados
    }
    
    with open(SALIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(datos_json, f, ensure_ascii=False, indent=2)
    
    tamano = os.path.getsize(SALIDA_JSON) / 1024 / 1024
    print(f"    Guardado: {SALIDA_JSON} ({tamano:.2f} MB)")
    
    # Generar Excel
    print(f"\n[5] Generando Excel...")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planos"
        
        headers = ["Carpeta", "Archivo", "Tabla", "Campo", "Valor"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="00A2D7", fill_type="solid")
        
        row = 2
        for plano in resultados:
            carpeta = plano.get('carpeta', '')
            archivo = plano.get('archivo', '')
            for i, tabla in enumerate(plano.get('tablas', []), 1):
                for campo, valor in tabla.get('atributos', {}).items():
                    ws.cell(row=row, column=1, value=carpeta)
                    ws.cell(row=row, column=2, value=archivo)
                    ws.cell(row=row, column=3, value=i)
                    ws.cell(row=row, column=4, value=campo)
                    ws.cell(row=row, column=5, value=valor)
                    row += 1
        
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 30
        
        wb.save(SALIDA_EXCEL)
        print(f"    Guardado: {SALIDA_EXCEL}")
        
    except Exception as e:
        print(f"    Error Excel: {e}")
    
    # Limpiar
    shutil.rmtree(temp_dir, ignore_errors=True)
    
    print(f"\n{'='*60}")
    print(f"FINALIZADO: {datetime.now()}")
    print("="*60)
    print(f"\nAhora abre visor_planos.html para ver los resultados")

if __name__ == "__main__":
    main()

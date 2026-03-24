# -*- coding: utf-8 -*-
"""
SISTEMA EXTRACTOR DE PLANOS - INTENTO 2 (ODA + ezdxf)
Convierte archivos DWG a DXF y extrae las tablas técnicas con atributos.
"""

import os
import sys
import json
import subprocess
import shutil
import tempfile
from datetime import datetime
from concurrent.futures import ProcessPoolExecutor, as_completed
import io

# Configurar UTF-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ================== CONFIGURACION ==================
SERVIDOR = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
SALIDA_JSON = r"output\datos_planos.json"
SALIDA_EXCEL = r"output\planos_tecnicos.xlsx"
CARPETA_DXF = r"\\192.168.2.37\ingenieria\PRODUCCION\_DXF_CONVERTIDOS"
ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"

# Importar ezdxf
try:
    import ezdxf
except ImportError:
    print("[ERROR] Instalar ezdxf: pip install ezdxf")
    sys.exit(1)

# ================== FUNCIONES ==================

def obtener_todos_dwg():
    """Busca todos los archivos DWG en el servidor"""
    print("[1] Buscando archivos DWG...")
    archivos = []
    
    for root, dirs, files in os.walk(SERVIDOR):
        # Ignorar carpeta de DXF convertidos
        if '_DXF_CONVERTIDOS' in root:
            continue
            
        for file in files:
            if file.lower().endswith('.dwg'):
                ruta_completa = os.path.join(root, file)
                archivos.append(ruta_completa)
    
    print(f"    Encontrados: {len(archivos)} archivos DWG")
    return archivos

def convertir_dwg(dwg_path, carpeta_salida, carpeta_entrada):
    """Convierte un DWG a DXF usando ODA"""
    try:
        cmd = [
            ODA_EXE,
            carpeta_entrada,
            carpeta_salida,
            "ACAD2018",
            "DXF",
            "0",  # Sin subdirectorios
            "0"
        ]
        
        resultado = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=600  # 10 min por carpeta
        )
        
        if resultado.returncode == 0:
            # Buscar el DXF generado
            nombre_dwg = os.path.basename(dwg_path)
            nombre_dxf = os.path.splitext(nombre_dwg)[0] + ".dxf"
            ruta_dxf = os.path.join(carpeta_salida, nombre_dxf)
            
            if os.path.exists(ruta_dxf):
                return ruta_dxf
            
            # Buscar en subcarpetas
            for root, dirs, files in os.walk(carpeta_salida):
                for f in files:
                    if f.lower().endswith('.dxf'):
                        return os.path.join(root, f)
        
        return None
        
    except Exception as e:
        print(f"    Error converting {dwg_path}: {e}")
        return None

def extraer_datos_dxf(ruta_dxf, nombre_original):
    """Extrae datos técnicos de un archivo DXF"""
    datos = {
        'archivo_original': nombre_original,
        'archivo_dxf': os.path.basename(ruta_dxf),
        'tablas': [],
        'error': None
    }
    
    try:
        doc = ezdxf.readfile(ruta_dxf)
        msp = doc.modelspace()
        
        # Buscar bloques con atributos (contienen las tablas técnicas)
        bloques = list(msp.query('INSERT'))
        
        tablas_encontradas = []
        
        for block_ref in bloques:
            if hasattr(block_ref, 'attribs') and block_ref.attribs:
                nombre_bloque = block_ref.dxf.name
                atributos = {}
                
                for attr in block_ref.attribs:
                    tag = getattr(attr.dxf, 'tag', '').strip() if hasattr(attr.dxf, 'tag') else ''
                    texto = getattr(attr.dxf, 'text', '').strip() if hasattr(attr.dxf, 'text') else ''
                    
                    if tag and texto and tag not in atributos:
                        atributos[tag] = texto
                
                if atributos:
                    # Crear identificador único para la tabla
                    # Usamos los primeros 2 atributos como clave
                    clave = "-".join(list(atributos.keys())[:2])
                    
                    # Verificar si ya existe una tabla con mismos datos
                    existe = False
                    for t in tablas_encontradas:
                        if t.get('atributos') == atributos:
                            existe = True
                            break
                    
                    if not existe:
                        tablas_encontradas.append({
                            'bloque': nombre_bloque,
                            'atributos': atributos
                        })
        
        datos['tablas'] = tablas_encontradas
        
        # También buscar texts que parezcan datos técnicos
        textos = []
        for ent in msp:
            if ent.dxftype() == 'TEXT':
                txt = getattr(ent, 'text', '').strip()
                if txt and len(txt) < 50 and any(c.isdigit() for c in txt):
                    textos.append(txt)
        
        datos['textos_sueltos'] = textos[:20]  # Limitar a 20
        
    except Exception as e:
        datos['error'] = str(e)
    
    return datos

def proceso_paralelo(args):
    """Procesa un archivo DWG (para usar con ProcessPoolExecutor)"""
    dwg_path, carpeta_dxf, carpeta_dwg = args
    
    nombre = os.path.basename(dwg_path)
    print(f"    Procesando: {nombre}")
    
    # Convertir a DXF
    dxf_path = convertir_dwg(dwg_path, carpeta_dxf, carpeta_dwg)
    
    if dxf_path and os.path.exists(dxf_path):
        # Extraer datos
        datos = extraer_datos_dxf(dxf_path, nombre)
        return datos
    else:
        return {
            'archivo_original': nombre,
            'error': 'No se pudo convertir a DXF',
            'tablas': []
        }

def procesar_todos():
    """Procesa todos los archivos DWG del servidor"""
    print("="*60)
    print("EXTRACTOR DE PLANOS TECNICOS - ODA + ezdxf")
    print("="*60)
    print(f"Servidor: {SERVIDOR}")
    print(f"Inicio: {datetime.now()}")
    
    # Obtener lista de archivos
    archivos_dwg = obtener_todos_dwg()
    
    if not archivos_dwg:
        print("[ERROR] No se encontraron archivos DWG")
        return
    
    # Crear carpeta para DXF si no existe
    if not os.path.exists(CARPETA_DXF):
        try:
            os.makedirs(CARPETA_DXF)
            print(f"[OK] Carpeta DXF creada: {CARPETA_DXF}")
        except Exception as e:
            print(f"[WARN] No se pudo crear carpeta DXF: {e}")
            CARPETA_DXF = tempfile.mkdtemp(prefix="dxf_cache_")
            print(f"    Usando temp: {CARPETA_DXF}")
    
    # Procesar archivos
    print(f"\n[2] Procesando {len(archivos_dwg)} archivos...")
    
    resultados = []
    errores = 0
    exitosos = 0
    
    # Procesar en paralelo (usar 75% de nucleos)
    import multiprocessing
    nucleos = max(1, int(multiprocessing.cpu_count() * 0.75))
    print(f"    Usando {nucleos} nucleos")
    
    # Preparar argumentos
    args_list = []
    for dwg in archivos_dwg:
        carpeta_dwg = os.path.dirname(dwg)
        args_list.append((dwg, CARPETA_DXF, carpeta_dwg))
    
    # Ejecutar en paralelo
    with ProcessPoolExecutor(max_workers=nucleos) as executor:
        futures = {executor.submit(proceso_paralelo, args): args[0] for args in args_list}
        
        for i, future in enumerate(as_completed(futures)):
            try:
                resultado = future.result()
                resultados.append(resultado)
                
                if resultado.get('error'):
                    errores += 1
                elif resultado.get('tablas'):
                    exitosos += 1
                    
                # Progreso cada 100 archivos
                if (i + 1) % 100 == 0:
                    print(f"    Progreso: {i+1}/{len(archivos_dwg)} (Error: {errores})")
                    
            except Exception as e:
                errores += 1
                print(f"    Error en proceso: {e}")
    
    print(f"\n[3] Resumen:")
    print(f"    Total procesados: {len(resultados)}")
    print(f"    Exitosos (con tablas): {exitosos}")
    print(f"    Errores: {errores}")
    
    # Guardar JSON
    print(f"\n[4] Guardando datos_planos.json...")
    datos_json = {
        'metadata': {
            'servidor': SERVIDOR,
            'total_archivos': len(resultados),
            'exitosos': exitosos,
            'errores': errores,
            'fecha_extraccion': datetime.now().isoformat()
        },
        'planos': resultados
    }
    
    os.makedirs(os.path.dirname(SALIDA_JSON), exist_ok=True)
    with open(SALIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(datos_json, f, ensure_ascii=False, indent=2)
    
    print(f"    Guardado: {SALIDA_JSON}")
    
    # Generar Excel
    print(f"\n[5] Generando Excel...")
    generar_excel(resultados)
    
    print(f"\n{'='*60}")
    print(f"FINALIZADO: {datetime.now()}")
    print("="*60)

def generar_excel(resultados):
    """Genera archivo Excel organizado por vehículo/carpeta"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="00A2D7", end_color="00A2D7", fill_type="solid")
        
        # Hoja principal
        ws = wb.active
        ws.title = "Planos Tecnicos"
        
        # Headers
        headers = ["Vehiculo", "Carpeta", "Archivo", "Tabla #", "Campo", "Valor"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Agregar datos
        row = 2
        
        for plano in resultados:
            archivo = plano.get('archivo_original', 'Unknown')
            tablas = plano.get('tablas', [])
            
            # Determinar vehiculo y carpeta
            if SERVIDOR in archivo:
                ruta_rel = archivo.replace(SERVIDOR, '').strip('\\')
            else:
                ruta_rel = archivo
            
            partes = ruta_rel.split('\\')
            vehiculo = partes[0] if len(partes) > 0 else 'Unknown'
            carpeta = '\\'.join(partes[1:-1]) if len(partes) > 1 else ''
            
            if tablas:
                for i, tabla in enumerate(tablas, 1):
                    atributos = tabla.get('atributos', {})
                    for campo, valor in atributos.items():
                        ws.cell(row=row, column=1, value=vehiculo)
                        ws.cell(row=row, column=2, value=carpeta)
                        ws.cell(row=row, column=3, value=archivo)
                        ws.cell(row=row, column=4, value=i)
                        ws.cell(row=row, column=5, value=campo)
                        ws.cell(row=row, column=6, value=valor)
                        row += 1
            else:
                # Sin tablas - solo registrar
                ws.cell(row=row, column=1, value=vehiculo)
                ws.cell(row=row, column=2, value=carpeta)
                ws.cell(row=row, column=3, value=archivo)
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value="SIN DATOS")
                ws.cell(row=row, column=6, value=plano.get('error', ''))
                row += 1
        
        # Ajustar anchos
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 25
        
        # Guardar
        os.makedirs(os.path.dirname(SALIDA_EXCEL), exist_ok=True)
        wb.save(SALIDA_EXCEL)
        print(f"    Guardado: {SALIDA_EXCEL}")
        
    except ImportError:
        print("    [WARN] openpyxl no instalado. Excel no generado.")
        print("    Instalar: pip install openpyxl")
    except Exception as e:
        print(f"    [ERROR] Generando Excel: {e}")

if __name__ == "__main__":
    procesar_todos()

# -*- coding: utf-8 -*-
"""
EXTRACTOR POR CARPETA - Procesa una carpeta específica
Genera Excel con hojas por subcarpeta
"""
import re
import os
import sys
import json
import subprocess
import shutil
import tempfile
from datetime import datetime

# Configurar UTF-8
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ================== CONFIGURACION ==================
SERVIDOR = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"
carpeta_nombre = "AUTOMATA"
# Importar ezdxf
try:
    import ezdxf
except ImportError:
    print("[ERROR] Instalar ezdxf: pip install ezdxf")
    sys.exit(1)

def es_pieza_090(nombre_archivo):
    """Devuelve True si el nombre del archivo corresponde a la pieza 090"""
    nombre_sin_extension = os.path.splitext(nombre_archivo)[0]
    return bool(re.search(r'(?<!\d)090(?!\d)', nombre_sin_extension))


def obtener_dwg_carpeta(carpeta_nombre):
    """Busca archivos DWG de pieza 090 en una carpeta específica del servidor"""
    ruta_carpeta = os.path.join(SERVIDOR, carpeta_nombre)
    
    if not os.path.exists(ruta_carpeta):
        print(f"[ERROR] La carpeta no existe: {ruta_carpeta}")
        return []
    
    print(f"[1] Buscando archivos DWG de pieza 090 en: {carpeta_nombre}")
    archivos = []
    
    contador_carpetas = 0

    for root, dirs, files in os.walk(ruta_carpeta):
        contador_carpetas += 1

        if contador_carpetas % 100 == 0:
            print(f"    Revisadas {contador_carpetas} carpetas...", flush=True)

        # Ignorar carpetas que no interesa recorrer
        dirs[:] = [d for d in dirs if d not in ['_DXF_CONVERTIDOS', 'DXF', 'output', '__pycache__']]

        if '_DXF_CONVERTIDOS' in root:
            continue
            
        for file in files:
            if file.lower().endswith('.dwg') and es_pieza_090(file):
                print(f"    [090] Encontrado: {file}", flush=True)
                ruta_completa = os.path.join(root, file)
                ruta_rel = ruta_completa.replace(ruta_carpeta, '').strip('\\')
                archivos.append({
                    'ruta_completa': ruta_completa,
                    'ruta_relativa': ruta_rel,
                    'carpeta': os.path.dirname(ruta_rel),
                    'nombre': file
                })
    
    print(f"    Encontrados: {len(archivos)} archivos DWG de pieza 090")
    return archivos

def obtener_dwg_carpeta(carpeta_nombre):
    """Busca archivos DWG en una carpeta específica del servidor"""
    ruta_carpeta = os.path.join(SERVIDOR, carpeta_nombre)
    
    if not os.path.exists(ruta_carpeta):
        print(f"[ERROR] La carpeta no existe: {ruta_carpeta}")
        return []
    
    print(f"[1] Buscando archivos en: {carpeta_nombre}")
    archivos = []
    
    for root, dirs, files in os.walk(ruta_carpeta):
        # Ignorar carpeta DXF
        if '_DXF_CONVERTIDOS' in root:
            continue
            
        for file in files:
            if file.lower().endswith('.dwg') and es_pieza_090(file):
                ruta_completa = os.path.join(root, file)
                ruta_rel = ruta_completa.replace(ruta_carpeta, '').strip('\\')
                archivos.append({
                    'ruta_completa': ruta_completa,
                    'ruta_relativa': ruta_rel,
                    'carpeta': os.path.dirname(ruta_rel),
                    'nombre': file
                })
    
    print(f"    Encontrados: {len(archivos)} archivos DWG")
    return archivos

def convertir_dwg(dwg_path, carpeta_salida):
    """Convierte un DWG a DXF usando ODA"""
    carpeta_entrada = os.path.dirname(dwg_path)
    nombre_dwg = os.path.basename(dwg_path)
    
    try:
        cmd = [
            ODA_EXE,
            carpeta_entrada,
            carpeta_salida,
            "ACAD2018",
            "DXF",
            "0",
            "0"
        ]
        
        resultado = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300
        )
        
        if resultado.returncode == 0:
            nombre_dxf = os.path.splitext(nombre_dwg)[0] + ".dxf"
            ruta_dxf = os.path.join(carpeta_salida, nombre_dxf)
            
            if os.path.exists(ruta_dxf):
                return ruta_dxf
            
            for f in os.listdir(carpeta_salida):
                if f.lower().endswith('.dxf'):
                    return os.path.join(carpeta_salida, f)
        
        return None
    except:
        return None

CAMPOS_TECNICOS = {
    "OFFSET": "OFFSET",
    "BN": "BN",
    "BN+D": "BN+D",
    "BN INT": "BN INT",
    "BANDA NEGRA": "BN",
    "ACERO": "ACERO",
    "STEEL": "ACERO",
    "OFFSET PC": "OFFSET PC",
}

def normalizar_texto(texto):
    if texto is None:
        return ""
    return " ".join(str(texto).replace("\\P", " ").split()).strip()

def canon_campo(texto):
    t = normalizar_texto(texto).upper()
    return CAMPOS_TECNICOS.get(t)

def agrupar_por_y(items, tolerancia=6.0):
    if not items:
        return []

    ordenados = sorted(items, key=lambda x: (-x["y"], x["x"]))
    grupos = []

    for item in ordenados:
        agregado = False
        for grupo in grupos:
            if abs(item["y"] - grupo[0]["y"]) <= tolerancia:
                grupo.append(item)
                agregado = True
                break
        if not agregado:
            grupos.append([item])

    for grupo in grupos:
        grupo.sort(key=lambda x: x["x"])

    return grupos

def extraer_tablas_insert(block_ref):
    nombre_bloque = block_ref.dxf.name
    items = []

    for attr in block_ref.attribs:
        tag = normalizar_texto(getattr(attr.dxf, "tag", ""))
        texto = normalizar_texto(getattr(attr.dxf, "text", ""))

        p = getattr(attr.dxf, "insert", None)
        x = float(p.x) if p else 0.0
        y = float(p.y) if p else 0.0

        if tag or texto:
            items.append({
                "tag": tag,
                "texto": texto,
                "x": x,
                "y": y,
            })

    filas = agrupar_por_y(items, tolerancia=6.0)
    tablas = []

    for fila in filas:
        atributos = {}
        i = 0

        while i < len(fila):
            actual = fila[i]
            campo = canon_campo(actual["texto"]) or canon_campo(actual["tag"])
            valor = None

            # Caso 1: etiqueta visual + valor a la derecha
            if canon_campo(actual["texto"]) and i + 1 < len(fila):
                siguiente = fila[i + 1]
                if not canon_campo(siguiente["texto"]):
                    valor = siguiente["texto"]
                    i += 2
                else:
                    i += 1

            # Caso 2: atributo directo, ej. tag=OFFSET texto=50
            elif canon_campo(actual["tag"]) and actual["texto"] and not canon_campo(actual["texto"]):
                valor = actual["texto"]
                i += 1

            else:
                i += 1

            if campo and valor and valor != campo:
                atributos[campo] = valor

        if atributos:
            tablas.append({
                "bloque": nombre_bloque,
                "atributos": atributos
            })

    return tablas

def extraer_datos(ruta_dxf, info_archivo):
    """Extrae datos técnicos del DXF"""
    datos = {
        'archivo': info_archivo['nombre'],
        'ruta': info_archivo['ruta_relativa'],
        'carpeta': info_archivo['carpeta'],
        'pieza': '090',
        'tablas': [],
        'resumen': [],
        'error': None
    }

    try:
        doc = ezdxf.readfile(ruta_dxf)
        msp = doc.modelspace()

        bloques = list(msp.query('INSERT'))

        for block_ref in bloques:
            if hasattr(block_ref, 'attribs') and block_ref.attribs:
                tablas_bloque = extraer_tablas_insert(block_ref)

                for tabla in tablas_bloque:
                    existe = False
                    for t in datos['tablas']:
                        if t.get('bloque') == tabla.get('bloque') and t.get('atributos') == tabla.get('atributos'):
                            existe = True
                            break

                    if not existe:
                        datos['tablas'].append(tabla)

                        for campo, valor in tabla.get('atributos', {}).items():
                            if campo in ['OFFSET', 'BN', 'BN+D', 'BN INT', 'ACERO', 'OFFSET PC']:
                                datos['resumen'].append(f"{campo}: {valor}")

    except Exception as e:
        datos['error'] = str(e)

    datos['resumen'] = list(dict.fromkeys(datos['resumen']))
    return datos

def agrupar_resultados_por_carpeta(resultados):
    """Agrupa los resultados por carpeta para usarlos en el visor HTML"""
    carpetas = {}

    for plano in resultados:
        carpeta = plano.get('carpeta', '')
        if carpeta == '':
            carpeta = 'Raiz'

        if carpeta not in carpetas:
            carpetas[carpeta] = {
                'nombre': carpeta,
                'archivos': []
            }

        carpetas[carpeta]['archivos'].append(plano)

    return list(carpetas.values())

def generar_excel_por_carpeta(resultados, carpeta_nombre):
    """Genera Excel con hojas por subcarpeta"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    # Archivo Excel
    excel_file = f"output/{carpeta_nombre}_planos.xlsx"
    
    wb = openpyxl.Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00A2D7", fill_type="solid")
    
    # Agrupar por subcarpeta
    por_carpeta = {}
    for plano in resultados:
        carpeta = plano.get('carpeta', 'ROOT')
        if carpeta == '':
            carpeta = 'Raiz'
        # Limpiar caracteres inválidos para Excel
        carpeta_limpia = carpeta.replace('\\', '_').replace('\\', '_').replace('/', '_').replace('\\', '_')
        if carpeta_limpia not in por_carpeta:
            por_carpeta[carpeta_limpia] = []
        por_carpeta[carpeta_limpia].append(plano)
    
    # Crear hoja por cada subcarpeta
    # Primero eliminar la hoja por defecto
    wb.remove(wb.active)
    
    for carpeta, planos in sorted(por_carpeta.items()):
        # Limitar a 31 caracteres y limpiar caracteres inválidos
        nombre_limpio = carpeta[:31].replace('\\', '_').replace('/', '_')
        ws = wb.create_sheet(title=nombre_limpio)
        
        # Headers
        headers = ["Archivo", "Tabla", "Campo", "Valor"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
        
        row = 2
        for plano in planos:
            archivo = plano.get('archivo', '')
            for i, tabla in enumerate(plano.get('tablas', []), 1):
                for campo, valor in tabla.get('atributos', {}).items():
                    ws.cell(row=row, column=1, value=archivo)
                    ws.cell(row=row, column=2, value=i)
                    ws.cell(row=row, column=3, value=campo)
                    ws.cell(row=row, column=4, value=valor)
                    row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 30
    
    # Guardar
    wb.save(excel_file)
    return excel_file

def main(carpeta_nombre):
    print("="*60)
    print(f"EXTRACTOR CARPETA: {carpeta_nombre}")
    print("="*60)
    print(f"Inicio: {datetime.now()}")
    print()
    print("Los archivos DWG originales NO se modifican")
    print()
    
    # Obtener archivos
    archivos = obtener_dwg_carpeta(carpeta_nombre)
    
    if not archivos:
        print(f"[ERROR] No hay archivos en {carpeta_nombre}")
        return
    
    # Carpeta temporal
    temp_dir = tempfile.mkdtemp(prefix="dxf_carpeta_")
    print(f"\n[2] Procesando {len(archivos)} archivos...")
    
    resultados = []
    exitosos = 0
    errores = 0
    
    for i, info in enumerate(archivos, 1):
        nombre = info['nombre']
        
        if i % 20 == 0 or i == 1:
            print(f"  [{i}/{len(archivos)}] {info['carpeta'][:40]}...", end=" ", flush=True)
        
        dxf_path = convertir_dwg(info['ruta_completa'], temp_dir)
        
        if dxf_path and os.path.exists(dxf_path):
            datos = extraer_datos(dxf_path, info)
            
            if datos.get('tablas'):
                exitosos += 1
                if i % 20 == 0:
                    print(f"OK")
            else:
                if i % 20 == 0:
                    print("Sin datos")
            
            resultados.append(datos)
        else:
            errores += 1
            if i % 20 == 0:
                print("Error")
            resultados.append({
                'archivo': nombre,
                'carpeta': info['carpeta'],
                'ruta': info['ruta_relativa'],
                'error': 'No se pudo convertir',
                'tablas': []
            })
        
        if i % 50 == 0:
            try:
                for f in os.listdir(temp_dir):
                    if f.endswith('.dxf'):
                        os.remove(os.path.join(temp_dir, f))
            except:
                pass
    
    print(f"\n[3] Resumen:")
    print(f"    Procesados: {len(resultados)}")
    print(f"    Con tablas: {exitosos}")
    print(f"    Errores: {errores}")
    
    # Generar Excel
    print(f"\n[4] Generando Excel...")
    excel_file = generar_excel_por_carpeta(resultados, carpeta_nombre)
    print(f"    Guardado: {excel_file}")
    
    # Guardar JSON también
    json_file = f"output/{carpeta_nombre}_datos.json"
    carpetas_agrupadas = agrupar_resultados_por_carpeta(resultados)

    datos_json = {
        'metadata': {
            'carpeta': carpeta_nombre,
            'servidor': SERVIDOR,
            'pieza_objetivo': '090',
            'total': len(resultados),
            'exitosos': exitosos,
            'errores': errores,
            'fecha': datetime.now().isoformat()
        },
        'carpetas': carpetas_agrupadas,
        'planos': resultados
    }
    
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(datos_json, f, ensure_ascii=False, indent=2)
    
    print(f"    Guardado: {json_file}")
    
    # Limpiar
    shutil.rmtree(temp_dir, ignore_errors=True)
    
    print(f"\n{'='*60}")
    print(f"FINALIZADO: {datetime.now()}")
    print("="*60)

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python extraer_carpeta.py NOMBRE_CARPETA")
        print("Ejemplo: python extraer_carpeta.py 00-Homologación_VZLA")
        sys.exit(1)
    
    carpeta = sys.argv[1]
    main(carpeta)

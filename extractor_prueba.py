# -*- coding: utf-8 -*-
"""
EXTRACTOR DE PRUEBA - Procesa 20 archivos para verificar funcionamiento
Guarda estructura de carpetas del servidor
"""

import os
import sys
import json
import subprocess
import shutil
import tempfile
from datetime import datetime

# Configurar UTF-8
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ================== CONFIGURACION ==================
SERVIDOR = r"\\192.168.2.37\ingenieria\PRODUCCION\AGP PLANOS TECNICOS"
SALIDA_JSON = r"output\datos_planos.json"
SALIDA_EXCEL = r"output\planos_tecnicos.xlsx"
CARPETA_DXF = r"\\192.168.2.37\ingenieria\PRODUCCION\_DXF_CONVERTIDOS"
ODA_EXE = r"C:\Program Files\ODA\ODAFileConverter 27.1.0\ODAFileConverter.exe"

# Importar ezdxf
try:
    import ezdxf
except ImportError:
    print("[ERROR] Instalar ezdxf: pip install ezdxf")
    sys.exit(1)

# ================== FUNCIONES ==================

def obtener_dwg_prueba(cantidad=20):
    """Busca archivos DWG para prueba"""
    print(f"[1] Buscando {cantidad} archivos DWG de prueba...")
    archivos = []
    
    for root, dirs, files in os.walk(SERVIDOR):
        if '_DXF_CONVERTIDOS' in root:
            continue
            
        for file in files:
            if file.lower().endswith('.dwg'):
                ruta_completa = os.path.join(root, file)
                # Calcular ruta relativa desde el servidor
                ruta_rel = ruta_completa.replace(SERVIDOR, '').strip('\\')
                archivos.append({
                    'ruta_completa': ruta_completa,
                    'ruta_relativa': ruta_rel,
                    'carpeta': os.path.dirname(ruta_rel),
                    'nombre': file
                })
        
        if len(archivos) >= cantidad:
            break
    
    print(f"    Encontrados: {len(archivos[:cantidad])} archivos para prueba")
    return archivos[:cantidad]

def convertir_dwg(dwg_path, carpeta_salida):
    """Convierte un DWG a DXF usando ODA"""
    carpeta_entrada = os.path.dirname(dwg_path)
    nombre_dwg = os.path.basename(dwg_path)
    
    try:
        cmd = [
            ODA_EXE,
            carpeta_entrada,
            carpeta_salida,
            "ACAD2018",
            "DXF",
            "0",
            "0"
        ]
        
        resultado = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=300  # 5 min
        )
        
        if resultado.returncode == 0:
            # Buscar DXF generado
            nombre_dxf = os.path.splitext(nombre_dwg)[0] + ".dxf"
            ruta_dxf = os.path.join(carpeta_salida, nombre_dxf)
            
            if os.path.exists(ruta_dxf):
                return ruta_dxf
            
            # Buscar cualquier DXF nuevo
            for f in os.listdir(carpeta_salida):
                if f.lower().endswith('.dxf'):
                    return os.path.join(carpeta_salida, f)
        
        return None
        
    except Exception as e:
        print(f"    Error: {e}")
        return None

def extraer_datos(ruta_dxf, info_archivo):
    """Extrae datos técnicos del DXF"""
    datos = {
        'archivo': info_archivo['nombre'],
        'ruta': info_archivo['ruta_relativa'],
        'carpeta': info_archivo['carpeta'],
        'tablas': [],
        'error': None
    }
    
    try:
        doc = ezdxf.readfile(ruta_dxf)
        msp = doc.modelspace()
        
        # Buscar bloques con atributos
        bloques = list(msp.query('INSERT'))
        
        for block_ref in bloques:
            if hasattr(block_ref, 'attribs') and block_ref.attribs:
                nombre_bloque = block_ref.dxf.name
                atributos = {}
                
                for attr in block_ref.attribs:
                    tag = getattr(attr.dxf, 'tag', '').strip() if hasattr(attr.dxf, 'tag') else ''
                    texto = getattr(attr.dxf, 'text', '').strip() if hasattr(attr.dxf, 'text') else ''
                    
                    if tag and texto and tag not in atributos:
                        atributos[tag] = texto
                
                if atributos:
                    # Verificar duplicados
                    existe = False
                    for t in datos['tablas']:
                        if t.get('atributos') == atributos:
                            existe = True
                            break
                    
                    if not existe:
                        datos['tablas'].append({
                            'bloque': nombre_bloque,
                            'atributos': atributos
                        })
        
    except Exception as e:
        datos['error'] = str(e)
    
    return datos

def main():
    print("="*60)
    print("EXTRACTOR DE PRUEBA - 20 archivos")
    print("="*60)
    print(f"Inicio: {datetime.now()}")
    
    # Obtener archivos de prueba
    archivos = obtener_dwg_prueba(20)
    
    if not archivos:
        print("[ERROR] No hay archivos DWG")
        return
    
    # Carpeta temporal para DXF
    temp_dir = tempfile.mkdtemp(prefix="dxf_prueba_")
    print(f"\n[2] Convirtiendo y extrayendo...")
    
    resultados = []
    exitosos = 0
    errores = 0
    
    for i, info in enumerate(archivos, 1):
        nombre = info['nombre']
        print(f"  [{i}/{len(archivos)}] {info['carpeta']}/{nombre}...", end=" ")
        
        # Convertir
        dxf_path = convertir_dwg(info['ruta_completa'], temp_dir)
        
        if dxf_path and os.path.exists(dxf_path):
            # Extraer
            datos = extraer_datos(dxf_path, info)
            
            if datos.get('tablas'):
                print(f"OK ({len(datos['tablas'])} tablas)")
                exitosos += 1
            else:
                print(f"Sin datos")
            
            resultados.append(datos)
        else:
            print(f"Error conversion")
            errores += 1
            resultados.append({
                'archivo': nombre,
                'carpeta': info['carpeta'],
                'ruta': info['ruta_relativa'],
                'error': 'No se pudo convertir',
                'tablas': []
            })
    
    # Resumen
    print(f"\n[3] Resumen:")
    print(f"    Procesados: {len(resultados)}")
    print(f"    Con tablas: {exitosos}")
    print(f"    Errores: {errores}")
    
    # Mostrar ejemplos de datos extraidos
    print(f"\n[4] Ejemplos de datos extraidos:")
    for r in resultados[:5]:
        if r.get('tablas'):
            print(f"  {r.get('carpeta', '')}/{r['archivo']}:")
            for t in r['tablas'][:2]:
                print(f"    {t['atributos']}")
    
    # Guardar JSON
    print(f"\n[5] Guardando JSON...")
    os.makedirs(os.path.dirname(SALIDA_JSON), exist_ok=True)
    
    datos_json = {
        'metadata': {
            'servidor': SERVIDOR,
            'total': len(resultados),
            'exitosos': exitosos,
            'errores': errores,
            'fecha': datetime.now().isoformat()
        },
        'planos': resultados
    }
    
    with open(SALIDA_JSON, 'w', encoding='utf-8') as f:
        json.dump(datos_json, f, ensure_ascii=False, indent=2)
    
    print(f"    Guardado: {SALIDA_JSON}")
    
    # Generar Excel
    print(f"\n[6] Generando Excel...")
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planos"
        
        # Headers
        headers = ["Carpeta", "Archivo", "Tabla", "Campo", "Valor"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="00A2D7", fill_type="solid")
        
        row = 2
        for plano in resultados:
            carpeta = plano.get('carpeta', '')
            archivo = plano.get('archivo', '')
            for i, tabla in enumerate(plano.get('tablas', []), 1):
                for campo, valor in tabla.get('atributos', {}).items():
                    ws.cell(row=row, column=1, value=carpeta)
                    ws.cell(row=row, column=2, value=archivo)
                    ws.cell(row=row, column=3, value=i)
                    ws.cell(row=row, column=4, value=campo)
                    ws.cell(row=row, column=5, value=valor)
                    row += 1
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 25
        
        wb.save(SALIDA_EXCEL)
        print(f"    Guardado: {SALIDA_EXCEL}")
        
    except Exception as e:
        print(f"    Error Excel: {e}")
    
    # Limpiar
    shutil.rmtree(temp_dir, ignore_errors=True)
    
    print(f"\n{'='*60}")
    print(f"FINALIZADO: {datetime.now()}")
    print("="*60)

if __name__ == "__main__":
    main()
